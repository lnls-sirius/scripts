#!/usr/bin/env python-sirius

import sys, time
from pickle import TRUE
from datetime import datetime, timedelta
from optparse import Values
from random import randint

import matplotlib.pyplot as plt
import numpy as np

from turtle import right

from openpyxl import Workbook
from openpyxl.compat import safe_string
from openpyxl.descriptors import Bool, MinMax, Min, Alias, NoneSet
from openpyxl.descriptors.serialisable import Serialisable
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color
from openpyxl.styles import colors
from openpyxl.cell import Cell

import epics

from siriuspy.clientarch import PVDataSet, Time


RAD_PVS = [
    'RAD:Thermo1:TotalDoseRate:Dose', 
    'RAD:Thermo2:TotalDoseRate:Dose', 
    'RAD:Thermo3:TotalDoseRate:Dose', 
    'RAD:Thermo4:TotalDoseRate:Dose', 
    'RAD:Thermo5:TotalDoseRate:Dose', 
    'RAD:Thermo6:TotalDoseRate:Dose', 
    'RAD:Thermo7:TotalDoseRate:Dose', 
    'RAD:Thermo8:TotalDoseRate:Dose', 
    'RAD:Thermo9:TotalDoseRate:Dose', 
    'RAD:Thermo10:TotalDoseRate:Dose', 
    'RAD:Thermo11:TotalDoseRate:Dose', 
    'RAD:Thermo12:TotalDoseRate:Dose', 
    'RAD:Thermo13:TotalDoseRate:Dose', 
    'RAD:Thermo14:TotalDoseRate:Dose',
    'RAD:Thermo15:TotalDoseRate:Dose',
    'RAD:Thermo16:TotalDoseRate:Dose',
    'RAD:Berthold:TotalDoseRate:Dose',
    'RAD:ELSE:TotalDoseRate:Dose',   
]


def create_worksheet():
    """Create worksheets."""

    # Preenchimento automático da tabela de injeção / cria a tabela e o arquivo .xls
    wb = Workbook()
    ws = wb.active

    # determina o tamanho das células das colunas 
    ws.column_dimensions['A'].width = 29.29
    ws.column_dimensions['B'].width = 8.50
    ws.column_dimensions['C'].width = 12.57

    # Alinhamento do Cabeçalho e cor
    ws['A2'].alignment = Alignment(horizontal="center", vertical="bottom")
    ws['B2'].alignment = Alignment(horizontal="center", vertical="bottom")
    ws['C2'].alignment = Alignment(horizontal="center", vertical="bottom")
    grayFill = PatternFill(start_color='FFD3D3D3',
                            end_color='FFD3D3D3',
                            fill_type='solid')
    ws['A2'].fill = grayFill
    ws['B2'].fill = grayFill
    ws['C2'].fill = grayFill

    # Cabeçalho da coluna 1
    ws['A2']='Descrição'
    ws['A2'].font = Font(size = 11, name = 'Calibri', bold=True)
    ws['A3']='Corrente Inicial = '
    ws['A4']='Corrente Final = '
    ws['A5']='Número de Pulsos na injeção =  '
    ws['A6']='Média de Corrente por pulso = '
    ws['A7']='Eficiência de Injeção SI = '
    ws['A8']='Eficiência de Injeção BO = '
    ws['A9']='Sintonia Bétatron Qx = '
    ws['A10']='Sintonia Bétatron Qy = '
    ws['A11']='Tensão Bias do Canhão = '
    ws['A12']='Tempo de vida inicial = '
    # ws['A13']='Dose inicial Thermo 15 = '
    # ws['A14']='Dose final Thermo 15 = '
    ws['A15']='Horário de início da injeção = '
    ws['A16']='Horário da liberação do feixe = '

    # Símbolos e grandezas da coluna 3
    ws['C2']='Símbolo'
    ws['C2'].font = Font(size = 11, name = 'Calibri', bold=True)
    ws['C3'] = '(mA)'
    ws['C4'] = '(mA)'
    ws['C5'] = ''
    ws['C6'] = '(mA)'
    ws['C7'] = '(% média)'
    ws['C8'] = '(% média)'
    ws['C9'] = ''
    ws['C10'] = ''
    ws['C11'] = '(V)'
    ws['C12'] = '(HH:mm:ss)'
    ws['C13'] = '(µSv)'
    ws['C14'] = '(µSv)'
    ws['C15'] = '(HH:mm)'
    ws['C16'] = '(HH:mm)'

    # Dados e cálculos
    ws['B2']='Dados'
    ws['B2'].font = Font(size = 11, name = 'Calibri', bold=True)

    return wb, ws


def set_border(ws, cell_range):
    """Função que define o estilo das bordas da tabela."""
    thin = Side(border_style="thin", color="FF000000")
    for row in ws[cell_range]:
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)


def set_font(ws, cell_range):
    """Função que define a fonte das células."""
    for row in ws[cell_range]:
        for cell in row:
            cell.font = Font(size = 11, name = 'Calibri', bold=False)


def set_alignment(ws, cell_range):
    """Função que define o alinhamento das células exceto cabeçalho."""
    for row in ws[cell_range]:
        for cell in row:
            cell.alignment = Alignment(horizontal = "right", vertical = "bottom")


def get_archiver_data(time_start, time_stop):
    """."""
    # Lista de PVs
    pvnames = [
        'SI-Glob:AP-CurrInfo:InjEff-Mon',
        'SI-Glob:AP-CurrInfo:Current-Mon',
        'AS-Glob:AP-CurrInfo:InjCount-Mon',
        'BO-Glob:AP-CurrInfo:RampEff-Mon',
        'LI-01:EG-BiasPS:voltoutsoft',
        'SI-Glob:DI-Tune-V:TuneFrac-Mon',
        'SI-13C4:DI-DCCT:Current-Mon',
        'SI-Glob:DI-Tune-H:TuneFrac-Mon',
        'SI-Glob:AP-CurrInfo:Lifetime-Mon',
        'SI-Glob:AP-CurrInfo:LifetimeBPM-Mon',
        'AS-Glob:AP-MachShift:Mode-Sts',
        'AS-RaMO:TI-EVG:InjectionEvt-Sel',
        ] + RAD_PVS
         
    # cria o objeto de consulta ao archiver
    pvdata = PVDataSet(pvnames)
    # determina intervalo de consulta, por exemplo, das 8h às 8h15 do dia 3/11/2022
    pvdata.time_start = time_start
    pvdata.time_stop = time_stop
    # consulta o archiverpvdata.update()
    pvdata.update()

    # for pvn in pvnames:
    #     print(pvn, pvdata[pvn].value)

    return pvdata


def fill_worksheet(time_start, time_stop):

    wb, ws = create_worksheet()
    set_border(ws, 'A2:C16')
    set_alignment(ws, 'C3:C16')
    set_font(ws, 'A3:C16')

    # pega dados do archiver
    pvdata = get_archiver_data(time_start, time_stop)
    pvdinjc = pvdata['AS-Glob:AP-CurrInfo:InjCount-Mon']

    # hora do inicio e fim da injeção 
    pvdmacs = pvdata['AS-Glob:AP-MachShift:Mode-Sts']
    timei = pvdmacs.value
    timef = pvdmacs.timestamp 
    for val, timestamp in zip(timei, timef):
        if val == 0:
            ws['B16']=datetime.fromtimestamp(timestamp).strftime("%H:%M:%S")
        if  val >= 3:
            ws['B15']=datetime.fromtimestamp(timestamp).strftime("%H:%M:%S")

    # calcula maxima variação de dosagem, e em qual sensor.
    dose_data = dict()
    dose_diff = dict()
    for pvname in RAD_PVS:
        pv_values = pvdata[pvname].value
        dose_data[pvname] = pv_values
        dose_diff[pvname] = pv_values[-1] - pv_values[0]
    result = list(dose_diff.values())
    maxdosediff = max(result)
    idx = result.index(maxdosediff)
    maxdose_pvname = RAD_PVS[idx]
    dose_ini = dose_data[maxdose_pvname][0]
    dose_end = dose_data[maxdose_pvname][-1]
    ws['B13']='{:.4f}'.format(dose_ini)  # Inicio RAD
    ws['B14']='{:.4f}'.format(dose_end)  # Final RAD
    maxdose_pvname = maxdose_pvname.replace(':TotalDoseRate:Dose', '').replace('RAD:', '')
    ws['A13']='Dose inicial {} ='.format(maxdose_pvname)
    ws['A14']='Dose final {} ='.format(maxdose_pvname)

    # cálculo para número de pulsos reias durante a injeção
    calci = pvdinjc.value[0]
    calcf = pvdinjc.value[-1]
    numpulsosreais = calcf - calci + 1

    # Loop para verificação da corrente no Anel
    # injeção ligada ou desligada
    pvd_injevt = pvdata['AS-RaMO:TI-EVG:InjectionEvt-Sel']
    pvd_current = pvdata['SI-Glob:AP-CurrInfo:Current-Mon']
    index = list(pvd_injevt.value).index(1)
    t_inj_on = pvd_injevt.timestamp[index]
    t_inj_off = pvd_injevt.timestamp[index+1]
    i_inj_on = np.interp(t_inj_on, pvd_current.timestamp, pvd_current.value)
    i_inj_off = np.interp(t_inj_off, pvd_current.timestamp, pvd_current.value)
    current_values = pvd_current.value
    if numpulsosreais >= 1:
        ws['B3'] = '{:.3f}'.format(current_values[0])
    if pvd_injevt.value[1] == True:
        ws['B4']='{:.3f}'.format(i_inj_off)
    mediaIA = (i_inj_off - i_inj_on) / numpulsosreais
    ws['B6']='{:.4f}'.format(mediaIA)  # Média da corrente injetada IA

  
    # Quantidade de pulsos calculada e eficiências de injeção BO e SI  
    ws['B5']='{}'.format(numpulsosreais)
    pvd_injeff = pvdata['SI-Glob:AP-CurrInfo:InjEff-Mon']
    ws['B7']='{:.3f}'.format(pvd_injeff.value.mean()) #EfSI
    pvd_rampeff = pvdata['BO-Glob:AP-CurrInfo:RampEff-Mon']
    ws['B8']='{:.3f}'.format(pvd_rampeff.value.mean())

    # valor da tensão de BIAS do Canhão
    pvd_bias = pvdata['LI-01:EG-BiasPS:voltoutsoft']
    ws['B11']='{:.2f}'.format(pvd_bias.value[-1])

    # sintonia X/Y
    pvd_tunev = pvdata['SI-Glob:DI-Tune-V:TuneFrac-Mon']
    ws['B10']='{:.4f}'.format(pvd_tunev.value[-1])
    pvd_tuneh = pvdata['SI-Glob:DI-Tune-H:TuneFrac-Mon']
    ws['B9']='{:.4f}'.format(pvd_tuneh.value[-1])

    # cálculo do tempo de vida do feixe
    pvd_lifetime = pvdata['SI-Glob:AP-CurrInfo:Lifetime-Mon']
    lt = pvd_lifetime.value[-1]
    ws['B12']=('{:d}:{:02d}:{:02d}'.format(int(lt // 3600), int((lt % 3600) // 60), int((lt % 3600) % 60)))

    # cria (grava) o arquivo
    ws.title = "ef_inj"
    wb.save("eficiência_de_injeção.xlsx")

time_start = Time(2022, 11, 25, 3, 56, 00)
time_stop = Time(2022, 11, 25, 4, 10, 30)
fill_worksheet(time_start, time_stop)