#!/usr/bin/env python-sirius

import sys, time
from pickle import TRUE
from datetime import datetime, timedelta
from optparse import Values
from random import randint
from datetime import date
import matplotlib.pyplot as plt
import numpy as np
from turtle import right
from openpyxl import Workbook
from openpyxl.compat import safe_string
from openpyxl.descriptors import Bool, MinMax, Min, Alias, NoneSet
from openpyxl.descriptors.serialisable import Serialisable
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color
from openpyxl.styles import colors
from openpyxl.cell import Cell
from PyQt5.QtWidgets import * 
from PyQt5.QtGui import * 
from PyQt5.QtCore import * 
from siriushla.sirius_application import SiriusApplication
from siriushla.widgets import SiriusMainWindow
import epics
from siriuspy.clientarch import PVDataSet, Time


RAD_PVS = [
    'RAD:Thermo1:TotalDoseRate:Dose', 
    'RAD:Thermo2:TotalDoseRate:Dose', 
    'RAD:Thermo3:TotalDoseRate:Dose', 
    'RAD:Thermo4:TotalDoseRate:Dose', 
    'RAD:Thermo5:TotalDoseRate:Dose', 
    'RAD:Thermo6:TotalDoseRate:Dose', 
    'RAD:Thermo7:TotalDoseRate:Dose', 
    'RAD:Thermo8:TotalDoseRate:Dose', 
    'RAD:Thermo9:TotalDoseRate:Dose', 
    'RAD:Thermo10:TotalDoseRate:Dose', 
    'RAD:Thermo11:TotalDoseRate:Dose', 
    'RAD:Thermo12:TotalDoseRate:Dose', 
    'RAD:Thermo13:TotalDoseRate:Dose', 
    'RAD:Thermo14:TotalDoseRate:Dose',
    'RAD:Thermo15:TotalDoseRate:Dose',
    'RAD:Thermo16:TotalDoseRate:Dose',
    'RAD:Berthold:TotalDoseRate:Dose'
]

datetime_ = datetime.now()
datetime_ = datetime_.strftime("%d/%m/%Y")
# print(datetime_)


def create_worksheet():
    """Create worksheets."""

    # Preenchimento automático da tabela de injeção / cria a tabela e o arquivo .xls
    wb = Workbook()
    ws = wb.active

    # determina o tamanho das células das colunas 
    ws.column_dimensions['A'].width = 29.29
    ws.column_dimensions['B'].width = 8.50
    ws.column_dimensions['C'].width = 12.57

    # Alinhamento do Cabeçalho e cor
    ws['A2'].alignment = Alignment(horizontal="center", vertical="bottom")
    ws['B2'].alignment = Alignment(horizontal="center", vertical="bottom")
    ws['C2'].alignment = Alignment(horizontal="center", vertical="bottom")
    grayFill = PatternFill(start_color='FFD3D3D3',
                            end_color='FFD3D3D3',
                            fill_type='solid')
    ws['A2'].fill = grayFill
    ws['B2'].fill = grayFill
    ws['C2'].fill = grayFill

    # Cabeçalho da coluna 1
    ws['A2']='Descrição'
    ws['A2'].font = Font(size = 11, name = 'Calibri', bold=True)
    ws['A3']='Horário de início da injeção '
    ws['A4']='Duração da injeção '
    ws['A5']='Corrente Inicial '
    ws['A6']='Corrente Final '
    ws['A7']='Tempo de vida inicial '
    ws['A8']='Número de Pulsos na injeção  '
    ws['A9']='Corrente injetada média por pulso '
    ws['A10']='Carga do E-Gun média por pulso '
    ws['A11']='Tensão Bias do E-Gun '
    ws['A12']='Eficiência média de Injeção SI '
    ws['A13']='Eficiência média de Rampa BO '
    ws['A14']='Sintonia Bétatron Qx '
    ws['A15']='Sintonia Bétatron Qy '
    ws['A18']='ID 06SB APU22 (CNB) Fase'
    ws['A19']='ID 07SP APU22 (CAT) Fase'
    ws['A20']='ID 08SB APU22 (EMA) Fase'
    ws['A21']='ID 09SA APU22 (MNC) Fase'
    ws['A22']='ID 11SP APU58 (IPE) Fase'
    ws['A23']='ID 10SB EPU50 (SAB)  Gap'
    ws['A24']='ID 10SB EPU50 (SAB) Fase'

    # Símbolos e grandezas da coluna 3
    ws['C2']='Unidade'
    ws['C2'].font = Font(size = 11, name = 'Calibri', bold=True)
    ws['C3'] = '    hh:mm:ss'
    ws['C4'] = '    mm:ss'
    ws['C5'] = '     mA'
    ws['C6'] = '     mA'
    ws['C7'] = '    hh:mm'
    ws['C8'] = '         '
    ws['C9'] = '     mA'
    ws['C10'] = '    nC'
    ws['C11'] = '     V'
    ws['C12'] = '     %'
    ws['C13'] = '     %'
    ws['C14'] = '           '
    ws['C15'] = '           '
    ws['C16'] = '     µSv'
    ws['C17'] = '     µSv'
    ws['C18'] = '    mm'
    ws['C19'] = '    mm'
    ws['C20'] = '    mm'
    ws['C21'] = '    mm'
    ws['C22'] = '    mm'
    ws['C23'] = '    mm'
    ws['C24'] = '    mm'
    

    # Dados e cálculos
    ws['B2']='Dados'
    ws['B2'].font = Font(size = 11, name = 'Calibri', bold=True)
    ws['A1'].font = Font(italic=True)
    ws['A1']=datetime_        
    

    return wb, ws


def set_border(ws, cell_range):
    """Função que define o estilo das bordas da tabela."""
    thin = Side(border_style="thin", color="FF000000")
    for row in ws[cell_range]:
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)


def set_font(ws, cell_range):
    """Função que define a fonte das células."""
    for row in ws[cell_range]:
        for cell in row:
            cell.font = Font(size = 11, name = 'Calibri', bold=False)


def set_alignment(ws, cell_range):
    """Função que define o alinhamento das célulasdas colunas B e C, exceto cabeçalho."""
    for row in ws[cell_range]:
        for cell in row:
            cell.alignment = Alignment(horizontal = "center", vertical = "bottom")


def set_alignment_a(ws, cell_range_a):
    """Função que define o alinhamento das células da coluna A, exceto cabeçalho."""
    for row in ws[cell_range_a]:
        for cell in row:
            cell.alignment = Alignment(horizontal = "right", vertical = "bottom")


def get_archiver_data(time_start, time_stop):
    """."""
    # Lista de PVs
    pvnames = [
        'SI-Glob:AP-CurrInfo:InjEff-Mon',
        'SI-Glob:AP-CurrInfo:Current-Mon',
        'AS-Glob:AP-CurrInfo:InjCount-Mon',
        'BO-Glob:AP-CurrInfo:RampEff-Mon',
        'LI-01:EG-BiasPS:voltoutsoft',
        'LI-01:DI-ICT-1:Charge-Mon',
        'SI-Glob:DI-Tune-V:TuneFrac-Mon',
        'SI-13C4:DI-DCCT:Current-Mon',
        'SI-Glob:DI-Tune-H:TuneFrac-Mon',
        'SI-Glob:AP-CurrInfo:Lifetime-Mon',
        'SI-Glob:AP-CurrInfo:LifetimeBPM-Mon',
        'AS-Glob:AP-MachShift:Mode-Sts',
        'AS-RaMO:TI-EVG:InjectionEvt-Sel',
        'AS-RaMO:TI-EVG:InjectionEvt-Sts',
        'RAD:ELSE:TotalDoseRate:Dose',
        'SI-06SB:ID-APU22:Phase-Mon',
        'SI-07SP:ID-APU22:Phase-Mon',
        'SI-08SB:ID-APU22:Phase-Mon',
        'SI-09SA:ID-APU22:Phase-Mon',
        'SI-11SP:ID-APU58:Phase-Mon',
        'SI-10SB:ID-EPU50:Gap-Mon',
        'SI-10SB:ID-EPU50:Phase-Mon',
        ] + RAD_PVS
         
    """cria o objeto de consulta ao archiver"""
    pvdata = PVDataSet(pvnames)
    """determina intervalo de consulta, por exemplo, das 8h às 8h15 do dia 3/11/2022"""
    pvdata.time_start = time_start
    pvdata.time_stop = time_stop
    pvdata.update()

    return pvdata


def fill_worksheet(time_start, time_stop):

    wb, ws = create_worksheet()
    set_border(ws, 'A2:C24')
    set_alignment(ws, 'B3:C24')
    set_alignment_a(ws, 'A3:B24')
    set_font(ws, 'A3:C24')

    # --- pega dados do archiver
    pvdata = get_archiver_data(time_start, time_stop)
    pvdinjc = pvdata['AS-Glob:AP-CurrInfo:InjCount-Mon']
    # deltai = pvdinjc.timestamp[-1]-pvdinjc.timestamp[0]
    
    
    # --- hora do inicio e fim da injeção
    pvdmacs = pvdata['AS-Glob:AP-MachShift:Mode-Sts']
    timei = pvdmacs.value
    timef = pvdmacs.timestamp
    
    for val, timestamp in zip(timei, timef):
        if val == 0:
            time_end = timestamp
        if  val == 3:
            ws['B3']=datetime.fromtimestamp(timestamp).strftime("%H:%M:%S")
            time_beg = timestamp
    dtime = time_end - time_beg
    dmin = int(dtime//60)
    dsec = int(round(dtime - 60 * dmin))
    ws['B4']=f'{dmin:02}:{dsec:02}'
    
    # ws['B25']=datetime.timestamp(deltatimei).strftime("%H:%M:%S")
    # lt = pvd_lifetime.value[-1]
    # ws['B25']=('{:d}:{:02d}'.format(int((deltatimei % 3600) // 60)))

    # --- calcula maxima variação de dosagem, e em qual sensor.

    dose_data = dict()
    dose_diff = dict()
    a = +1
    b = 0
    for pvname in RAD_PVS:
        pv_values = pvdata[pvname].value
        dose_data[pvname] = pv_values
        dose_diff[pvname] = pv_values[-1] - pv_values[0]
        c = b + a
    result = list(dose_diff.values())
    maxdosediff = max(result)
    idx = result.index(maxdosediff)
    maxdose_pvname = RAD_PVS[idx]
    dose_ini = dose_data[maxdose_pvname][0]
    dose_end = dose_data[maxdose_pvname][-1]
    ws['B16']='{:.3f}'.format(dose_ini)  # Inicio RAD
    ws['B17']='{:.3f}'.format(dose_end)  # Final RAD
    maxdose_pvname = maxdose_pvname.replace(':TotalDoseRate:Dose', '').replace('RAD:', '')
    ws['A16']='Dose inicial {} '.format(maxdose_pvname)
    ws['A17']='Dose final {} '.format(maxdose_pvname)
    # cálculo para número de pulsos reias durante a injeção
    calci = pvdinjc.value[0]
    calcf = pvdinjc.value[-1]
    numpulsosreais = calcf - calci + 1

    # --- Loop para verificação da corrente no Anel

    # --- injeção ligada ou desligada"""
    pvd_injevt = pvdata['AS-RaMO:TI-EVG:InjectionEvt-Sts']
    pvd_current = pvdata['SI-Glob:AP-CurrInfo:Current-Mon']
    index = list(pvd_injevt.value).index(1)
    t_inj_on = pvd_injevt.timestamp[index]
    t_inj_off = pvd_injevt.timestamp[index+1]
    i_inj_on = np.interp(t_inj_on, pvd_current.timestamp, pvd_current.value)
    i_inj_off = np.interp(t_inj_off+1, pvd_current.timestamp, pvd_current.value)
    current_values = pvd_current.value
    if numpulsosreais >= 1:
        ws['B5'] = '{:.1f}'.format(current_values[0])
    if pvd_injevt.value[1] == True:
        ws['B6']='{:.1f}'.format(i_inj_off)
    mediaIA = (i_inj_off - i_inj_on) / numpulsosreais
    ws['B9']='{:.2f}'.format(mediaIA)  #Média da corrente injetada IA"""
    
  
    # --- Quantidade de pulsos calculada e eficiências de injeção BO e SI
    ws['B8']='{}'.format(numpulsosreais)
    pvd_injeff = pvdata['SI-Glob:AP-CurrInfo:InjEff-Mon']
    ws['B12']='{:.1f}'.format(pvd_injeff.value.mean()) #EfSI
    pvd_rampeff = pvdata['BO-Glob:AP-CurrInfo:RampEff-Mon']
    ws['B13']='{:.1f}'.format(pvd_rampeff.value.mean())

    # --- valor da tensão de BIAS do Canhão
    pvd_bias = pvdata['LI-01:EG-BiasPS:voltoutsoft']
    ws['B11']='{:.2f}'.format(pvd_bias.value[-1])
    

    # --- sintonia X/Y
    pvd_tunev = pvdata['SI-Glob:DI-Tune-V:TuneFrac-Mon']
    ws['B15']='{:.3f}'.format(pvd_tunev.value[-1])
    pvd_tuneh = pvdata['SI-Glob:DI-Tune-H:TuneFrac-Mon']
    ws['B14']='{:.3f}'.format(pvd_tuneh.value[-1])

    # --- cálculo do tempo de vida do feixe
    pvd_lifetime = pvdata['SI-Glob:AP-CurrInfo:Lifetime-Mon']
    lt = pvd_lifetime.value[-1]
    ws['B7']=('{:d}:{:02d}'.format(int(lt // 3600), int((lt % 3600) // 60)))
    
    # --- valores de Phase e Gap dos IDs
    pvidcar_phase = pvdata['SI-06SB:ID-APU22:Phase-Mon']
    ws['B18']='{:.3f}'.format(pvidcar_phase.value[-1])
    pvidcat_phase = pvdata['SI-07SP:ID-APU22:Phase-Mon']
    ws['B19']='{:.3f}'.format(pvidcat_phase.value[-1])
    pvidema_phase = pvdata['SI-08SB:ID-APU22:Phase-Mon']
    ws['B20']='{:.3f}'.format(pvidema_phase.value[-1])
    pvidman_phase = pvdata['SI-09SA:ID-APU22:Phase-Mon']
    ws['B21']='{:.3f}'.format(pvidman_phase.value[-1])
    pvidipe_phase = pvdata['SI-11SP:ID-APU58:Phase-Mon']
    ws['B22']='{:.3f}'.format(pvidipe_phase.value[-1])
    pvidisab_gap = pvdata['SI-10SB:ID-EPU50:Gap-Mon']
    ws['B23']='{:.3f}'.format(pvidisab_gap.value[-1])
    pvidisab_phase = pvdata['SI-10SB:ID-EPU50:Phase-Mon']
    ws['B24']='{:.3f}'.format(pvidisab_phase.value[-1])

    # --- valor de média da carga do E-Gun na injeção
    pvdlicharge = pvdata['LI-01:DI-ICT-1:Charge-Mon']
    ws['B10']='{:.3f}'.format(pvdlicharge.value.mean())
    
    
    # com segundos ==> ws['B12']=('{:d}:{:02d}:{:02d}'.format(int(lt // 3600), int((lt % 3600) // 60), int((lt % 3600) % 60)))
    # cria (grava) o arquivo
    ws.title = "ef_inj"
    # wb.save("/home/sirius/repos-dev/scripts/bin/eficiência_de_injeção.xlsx")
    wb.save("eficiência_de_injeção.xlsx")


class Window(SiriusMainWindow): 
    """."""

    def __init__(self, parent=None): 
        super().__init__() 
        self.setWindowTitle("Python ") 
        self.setGeometry(100, 100, 310, 200) 
        self.UiComponents()
        self.show()
        
    def UiComponents(self): 

        button = QPushButton("Gerar Tabela", self) 
        button.setGeometry(200, 150, 100, 30)
        button.setStyleSheet('background-color: red')
        button.setStyleSheet('background-color:#ff0010;')
        button.setStyleSheet('background-color:rgb(200,0,0)')
        button.move(110, 160)
        button.pressed.connect(self.clickme)
        
        today = date.today()

        self.tb_yyi = QLineEdit(self)
        self.tb_yyi.move(10, 15)
        self.tb_yyi.resize(40,20)
        self.tb_yyi.setText(str(today.year))
        self.tb_mi = QLineEdit(self)
        self.tb_mi.move(60, 15) 
        self.tb_mi.resize(40,20)
        self.tb_mi.setText(str(today.month))
        self.tb_di = QLineEdit(self)
        self.tb_di.move(110, 15) 
        self.tb_di.resize(40,20)
        self.tb_di.setText(str(today.day))
        self.tb_hhi = QLineEdit(self)
        self.tb_hhi.move(160, 15)
        self.tb_hhi.resize(40,20)
        self.tb_mmi = QLineEdit(self)
        self.tb_mmi.move(210, 15) 
        self.tb_mmi.resize(40,20)
        self.tb_ssi = QLineEdit(self)
        self.tb_ssi.move(260, 15) 
        self.tb_ssi.resize(40,20)
        self.tb_ssi.setText(str(0))

        self.tb_interval = QLineEdit(self)
        self.tb_interval.move(10, 95)
        self.tb_interval.resize(40,20)
        self.tb_interval.setText('15')
        # self.tb_yyf = QLineEdit(self)
        # self.tb_mf = QLineEdit(self)
        # self.tb_mf.move(60, 95) 
        # self.tb_mf.resize(40,20)
        # self.tb_df = QLineEdit(self)
        # self.tb_df.move(110, 95) 
        # self.tb_df.resize(40,20)
        # self.tb_hhf = QLineEdit(self)
        # self.tb_hhf.move(160, 95)
        # self.tb_hhf.resize(40,20)
        # self.tb_mmf = QLineEdit(self)
        # self.tb_mmf.move(210, 95) 
        # self.tb_mmf.resize(40,20)
        # self.tb_ssf = QLineEdit(self)
        # self.tb_ssf.move(260, 95) 
        # self.tb_ssf.resize(40,20)
        
        self.lay_i = QLabel(self)
        self.lay_i.move (15, 5)
        self.lay_i.resize (260, 100)
        self.lay_i.setText('Data/hora inicial - yyyy:mm:dd:hh:mm:ss')
        self.lay_f = QLabel(self)
        self.lay_f.move (15, 80)
        self.lay_f.resize (260, 100)
        self.lay_f.setText('Minutos de intervalo')
        self.lay_p = QLabel(self)
        self.lay_p.move (15, 115)
        self.lay_p.resize (260, 100)
        
        # self.lay_p.setText('Tabela Ok!')

        wid = QWidget()
        lay = QGridLayout(wid)
        lay.setAlignment(Qt.AlignCenter)
        lay.addWidget(self.tb_yyi, 0, 0)
        lay.addWidget(self.tb_mi, 0, 1)
        lay.addWidget(self.tb_di, 0, 2)
        lay.addWidget(self.tb_hhi, 0, 3)
        lay.addWidget(self.tb_mmi, 0, 4)
        lay.addWidget(self.tb_ssi, 0, 5)
        lay.addWidget(self.lay_i, 1, 0, 1, 0)
        lay.addWidget(self.tb_interval, 2, 0)
        # lay.addWidget(self.tb_mf, 2, 1)
        # lay.addWidget(self.tb_df, 2, 2)
        # lay.addWidget(self.tb_hhf, 2, 3)
        # lay.addWidget(self.tb_mmf, 2, 4)
        # lay.addWidget(self.tb_ssf, 2, 5)
        lay.addWidget(self.lay_f, 3, 0, 3, 0)
        lay.addWidget(button, 6, 0, 6, 0)
        self.setCentralWidget(wid)
        return wid

        # lay.addWidget(lbl_corr, 0, 0)
        # lay.addWidget(led_corr, 0, 1)
        # lay.addWidget(sts_corr, 0, 2)
        # lay.addWidget(cnf_corr, 0, 3)
        # lay.addWidget(lbl_ctrl, 1, 0)
        # lay.addWidget(led_ctrl, 1, 1)
        # lay.addWidget(sts_ctrl, 1, 2)

    def clickme(self): 
        self.lay_p.setText('TABELA OK!')
        year = int(self.tb_yyi.text())
        month = int(self.tb_mi.text())
        day = int(self.tb_di.text())
        hour = int(self.tb_hhi.text())
        minute = int(self.tb_mmi.text())
        second = int(self.tb_ssi.text())
        interval = int(self.tb_interval.text())
        time_start= Time(year, month, day, hour, minute, second)
        time_stop = time_start + 60 * interval
        fill_worksheet(time_start, time_stop)
        print('Ok!', day,"/",month,"/",year)
        
        # datetime_
        

app = SiriusApplication()
app.open_window(Window)
sys.exit(app.exec()) 